"""Privacy-safe, bounded extraction of statement spreadsheets (.xlsx).

Unlike a PDF, a spreadsheet is already structured, so extraction reads the
cells into a bounded in-memory grid rather than pulling free text. Raw upload
bytes are written to a server-owned temp file and deleted on every exit.
"""

from __future__ import annotations

import asyncio
import hashlib
import logging
import tempfile
from collections.abc import AsyncIterator, Iterator
from contextlib import asynccontextmanager, contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.importing.document import (
    AsyncReadable,
    DocumentLimits,
    _write_chunk,
    sanitize_filename,
)
from app.importing.errors import UnsupportedDocumentError

# xlsx is a ZIP container; every valid file starts with the local-file magic.
ZIP_MAGIC = b"PK\x03\x04"
XLSX_MIME_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
        "",
    }
)
_MAX_SHEETS = 12
_MAX_ROWS_PER_SHEET = 10_000
_MAX_COLS = 24


@dataclass(frozen=True, slots=True, repr=False)
class ExtractedWorkbook:
    """Ephemeral, bounded cell grid extracted from one validated spreadsheet."""

    document_id: str
    sha256: str
    sanitized_filename: str
    byte_count: int
    sheet_count: int
    sheets: dict[str, tuple[tuple[str, ...], ...]] = field(repr=False)

    def sheet(self, name: str) -> tuple[tuple[str, ...], ...] | None:
        """Return a sheet's rows by exact name, or ``None`` if absent."""

        return self.sheets.get(name)


def _validate_boundary(filename: str, content_type: str) -> str:
    safe_name = sanitize_filename(filename)
    if Path(safe_name).suffix.casefold() != ".xlsx":
        raise UnsupportedDocumentError("statement must use the .xlsx extension")
    normalized_mime = content_type.partition(";")[0].strip().casefold()
    if normalized_mime not in XLSX_MIME_TYPES:
        raise UnsupportedDocumentError(
            "statement content type must be an Excel .xlsx workbook"
        )
    return safe_name


def _validate_magic(path: Path) -> None:
    with path.open("rb") as stream:
        if stream.read(len(ZIP_MAGIC)) != ZIP_MAGIC:
            raise UnsupportedDocumentError("statement content is not an .xlsx workbook")


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # Avoid trailing-zero noise while keeping exact currency magnitudes.
        return repr(value)
    return str(value).strip()


def _extract_workbook(
    path: Path,
    *,
    document_id: str,
    sha256: str,
    sanitized_filename: str,
    byte_count: int,
) -> ExtractedWorkbook:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError, ValueError):
        raise UnsupportedDocumentError(
            "statement workbook is malformed or unreadable"
        ) from None
    try:
        names = workbook.sheetnames[:_MAX_SHEETS]
        sheets: dict[str, tuple[tuple[str, ...], ...]] = {}
        for name in names:
            worksheet = workbook[name]
            rows: list[tuple[str, ...]] = []
            for index, row in enumerate(
                worksheet.iter_rows(max_col=_MAX_COLS, values_only=True)
            ):
                if index >= _MAX_ROWS_PER_SHEET:
                    break
                rows.append(tuple(_cell_text(cell) for cell in row))
            sheets[name] = tuple(rows)
    finally:
        workbook.close()
    if not sheets:
        raise UnsupportedDocumentError("statement workbook has no readable sheets")
    return ExtractedWorkbook(
        document_id=document_id,
        sha256=sha256,
        sanitized_filename=sanitized_filename,
        byte_count=byte_count,
        sheet_count=len(sheets),
        sheets=sheets,
    )


def _stage(
    path: Path,
    *,
    filename: str,
    content_type: str,
    limits: DocumentLimits,
    document_id: str,
    digest: hashlib._Hash,
    byte_count: int,
) -> ExtractedWorkbook:
    safe_name = _validate_boundary(filename, content_type)
    _validate_magic(path)
    return _extract_workbook(
        path,
        document_id=document_id,
        sha256=digest.hexdigest(),
        sanitized_filename=safe_name,
        byte_count=byte_count,
    )


@contextmanager
def stage_spreadsheet(
    stream: BinaryIO,
    *,
    filename: str,
    content_type: str,
    limits: DocumentLimits | None = None,
    temp_root: Path | None = None,
    logger: logging.Logger | None = None,
) -> Iterator[ExtractedWorkbook]:
    """Stage and extract a synchronous .xlsx upload, deleting bytes on exit."""

    effective_limits = limits or DocumentLimits()
    # Validate the boundary before reading bytes so bad uploads fail fast.
    _validate_boundary(filename, content_type)
    document_id = uuid4().hex
    with tempfile.TemporaryDirectory(prefix="statement-", dir=temp_root) as directory:
        path = Path(directory) / "upload.xlsx"
        digest = hashlib.sha256()
        byte_count = 0
        with path.open("wb") as destination:
            while chunk := stream.read(effective_limits.chunk_size):
                byte_count = _write_chunk(
                    destination,
                    chunk,
                    digest=digest,
                    byte_count=byte_count,
                    limits=effective_limits,
                )
        workbook = _stage(
            path,
            filename=filename,
            content_type=content_type,
            limits=effective_limits,
            document_id=document_id,
            digest=digest,
            byte_count=byte_count,
        )
        if logger is not None:
            logger.info(
                "statement_workbook_staged",
                extra={
                    "document_id": document_id,
                    "byte_count": byte_count,
                    "sheet_count": workbook.sheet_count,
                },
            )
        yield workbook


@asynccontextmanager
async def stage_spreadsheet_async(
    stream: AsyncReadable,
    *,
    filename: str,
    content_type: str,
    limits: DocumentLimits | None = None,
    temp_root: Path | None = None,
    logger: logging.Logger | None = None,
) -> AsyncIterator[ExtractedWorkbook]:
    """Stage an async .xlsx upload with cleanup on success, failure, or cancel."""

    effective_limits = limits or DocumentLimits()
    _validate_boundary(filename, content_type)
    document_id = uuid4().hex
    with tempfile.TemporaryDirectory(prefix="statement-", dir=temp_root) as directory:
        path = Path(directory) / "upload.xlsx"
        digest = hashlib.sha256()
        byte_count = 0
        try:
            with path.open("wb") as destination:
                while chunk := await stream.read(effective_limits.chunk_size):
                    byte_count = _write_chunk(
                        destination,
                        chunk,
                        digest=digest,
                        byte_count=byte_count,
                        limits=effective_limits,
                    )
        except asyncio.CancelledError:
            raise
        workbook = _stage(
            path,
            filename=filename,
            content_type=content_type,
            limits=effective_limits,
            document_id=document_id,
            digest=digest,
            byte_count=byte_count,
        )
        if logger is not None:
            logger.info(
                "statement_workbook_staged",
                extra={
                    "document_id": document_id,
                    "byte_count": byte_count,
                    "sheet_count": workbook.sheet_count,
                },
            )
        yield workbook
